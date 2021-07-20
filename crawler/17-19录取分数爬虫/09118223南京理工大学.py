# -*- coding: utf-8 -*-
"""
Created on Mon May 18 16:04:40 2020

@author: lenovo
"""

# -*- coding: utf-8 -*-
"""
Created on Sun May 17 18:22:38 2020

@author: lenovo
"""

import requests
import random
from  openpyxl import  Workbook
from urllib.request import quote
import json
 
USER_AGENTS = [#将爬虫伪装成浏览器，防止反爬虫
        "Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1"
        "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
        "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
        "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
        "Opera/9.80 (Windows NT 5.1; U; zh-cn) Presto/2.9.168 Version/11.50",
        "Mozilla/5.0 (Windows NT 5.1; rv:5.0) Gecko/20100101 Firefox/5.0",
        "Mozilla/5.0 (Windows NT 5.2) AppleWebKit/534.30 (KHTML, like Gecko) Chrome/12.0.742.122 Safari/534.30",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; 360SE)",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0",
        "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.2)",
        "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
        "Mozilla/4.0 (compatible; MSIE 5.0; Windows NT)",
        "Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070309 Firefox/2.0.0.3",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12 "
        ]

headers = {#配置headers
    'User-agent': random.choice(USER_AGENTS), #设置get请求的User-Agent，用于伪装浏览器UA 
    'Connection': 'keep-alive',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Host': 'zsb.njust.edu.cn',
    'Referer': 'http://zsb.njust.edu.cn/lqjh_fsx'
}

wb=Workbook()#创建工作簿
index=1#初始化写入Excel表格的行数

for sheet in wb.sheetnames:
    wb.remove(wb[sheet])#清除初始的第一个sheet
    
sheet = wb.create_sheet(title='各专业历年录取分数线')#命名生成的sheet

sheet.column_dimensions['A'].width = 12#设置表格各列的宽度以及表头
sheet.column_dimensions['B'].width = 7
sheet.column_dimensions['E'].width = 90
sheet.column_dimensions['F'].width = 7
sheet.column_dimensions['G'].width = 15
sheet.freeze_panes = 'A1'
sheet['A1'] = 'College'
sheet['B1'] = 'Year'
sheet['C1'] = 'Province'
sheet['D1'] = 'Category'
sheet['E1'] = 'Major'
sheet['F1'] = 'Score'
sheet['G1'] = 'Contributor'

url1 = 'http://zsb.njust.edu.cn/lqScore/initDateWebCon?pageSize=15&rowoffset=0&val1='#基础网址
url2='http://zsb.njust.edu.cn/lqPain/initDateCon?pageSize=15&rowoffset=0&val1=&val2=&val3='
provinceset=['内蒙古','黑龙江','吉林','辽宁','北京','天津','河北','河南','山西','山东','陕西','宁夏','甘肃','青海','新疆','西藏','四川','重庆','云南','贵州','广东','广西','湖南','湖北','江西','福建','浙江','上海','安徽','江苏','海南']#省份列表

for province in provinceset:
    line1=url1+province#拼接成真实网址
    line2=url2+province
    line1=quote(line1, safe=";/?:@&=+$,", encoding="utf-8")#对网址编码
    line2=quote(line2, safe=";/?:@&=+$,", encoding="utf-8")
    req1=requests.request("GET",line1,headers=headers)#爬取对应网页内容
    req2=requests.request("GET",line2,headers=headers)
    text1=json.loads(req1.text)#转变成字典格式
    text2=json.loads(req2.text)
    content1=text1['data']['list']#取出包含分数线的字典
    content2=text2['data']['list']
    for item1 in content1:
        if item1['class_name']=='本科一批':#只处理正常的本科一批
            major=item1['professional_name']#文理和分数线分别在两个网址之中，所以遍历另一个网址找出文理科分类
            for item2 in content2:
                if item2['professional_name']==major:
                    category=item2['subject']
                    break
            if item1['year1']!=None:#确保当年分数是有的
                index=index+1
                indexstr=str(index)
                sheet['A'+indexstr] = '南京理工大学'#写入数据
                sheet['B'+indexstr]='2017'
                sheet['C'+indexstr]=item1['province']
                sheet['D'+indexstr]=category
                sheet['E'+indexstr]=major
                sheet['F'+indexstr]=item1['year1']
                sheet['G'+indexstr]='09118223吴亦珂'
            if item1['year2']!=None:#确保当年分数是有的
                index=index+1
                indexstr=str(index)
                sheet['A'+indexstr] = '南京理工大学'#写入数据
                sheet['B'+indexstr]='2018'
                sheet['C'+indexstr]=item1['province']
                sheet['D'+indexstr]=category
                sheet['E'+indexstr]=major
                sheet['F'+indexstr]=item1['year2']
                sheet['G'+indexstr]='09118223吴亦珂'
            if item1['year3']!=None:#确保当年分数是有的
                index=index+1
                indexstr=str(index)
                sheet['A'+indexstr] = '南京理工大学'#写入数据
                sheet['B'+indexstr]='2019'
                sheet['C'+indexstr]=item1['province']
                sheet['D'+indexstr]=category
                sheet['E'+indexstr]=major
                sheet['F'+indexstr]=item1['year3']
                sheet['G'+indexstr]='09118223吴亦珂'
for i in range(2,index+1):#数据后续处理，如遇到多个分数的取最高分
    tempscore=sheet.cell(i,6).value
    if '/' in tempscore:
        scorelist=tempscore.split('/')
        sheet.cell(i,6).value=max(scorelist)
wb.save('09118223吴亦珂-南京理工大学.csv')#保存数据
